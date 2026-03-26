from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "plantilla-turnos-pazy.xlsx"
DRIVE_FOLDER_ID = "1QqodtcRBgpMYJEyJpM454svceGUWpgK3"


HEAD_FILL = PatternFill("solid", fgColor="7A0A3F")  # Pazy granate (placeholder)
HEAD_FONT = Font(color="FFFFFF", bold=True)
HEAD_ALIGN = Alignment(horizontal="center", vertical="center")


def style_header(ws, row=1):
    ws.freeze_panes = "A2"
    ws.row_dimensions[row].height = 20
    for cell in ws[row]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = HEAD_ALIGN


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 45)


def main():
    wb = Workbook()

    # Comerciales
    ws = wb.active
    ws.title = "Comerciales"
    ws.append(["id", "nombre", "activo"])
    style_header(ws)
    for i, name in enumerate(
        [
            "Georgi Valeriev",
            "Magüi Cerdá",
            "Antonella Sipan",
            "Iñigo Puyol",
            "Luz Romero",
            "Patricia López",
            "Jorge Romera",
            "Irene Peñalosa",
            "Maria Jose Rubio",
            "Alessandra Solis",
            "Adrian Garces",
            "Ignacio Rivas",
            "Alonso García",
            "Rodrigo Fernandez",
            "Lara Carrasco",
        ],
        start=1,
    ):
        ws.append([f"C{i:03d}", name, True])
    autosize(ws)

    # Vacaciones
    ws = wb.create_sheet("Vacaciones")
    ws.append(["nombre", "desde", "hasta", "motivo"])
    style_header(ws)
    ws.append(["Ejemplo Persona", date.today().isoformat(), date.today().isoformat(), "Vacaciones"])
    autosize(ws)

    # Turnos
    ws = wb.create_sheet("Turnos")
    ws.append(["weekStart", "fecha", "franja", "tipo", "modo", "asignadoA", "nota"])
    style_header(ws)
    # Ejemplo mínimo (la app rellenará 7*3*2 filas por semana)
    ws.append(
        [
            "2026-01-01",
            "2026-01-01",
            "MANANA",
            "FIJO",
            "NORMAL",
            "Georgi Valeriev",
            "",
        ]
    )
    autosize(ws)

    # Cambios
    ws = wb.create_sheet("Cambios")
    ws.append(["timestamp", "weekStart", "slotId", "antes", "despues", "motivo", "autor"])
    style_header(ws)
    ws.append([None, "2026-01-01", "2026-01-01__MANANA__FIJO", "A", "B", "Cambio acordado", ""])
    autosize(ws)

    # Config
    ws = wb.create_sheet("Config")
    ws.append(["clave", "valor"])
    style_header(ws)
    ws.append(["carpetaDriveId", DRIVE_FOLDER_ID])
    ws.append(["zonaHoraria", "Europe/Madrid"])
    ws.append(["permitirTodos", "TRUE"])
    autosize(ws)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()

